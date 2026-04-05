"""
COMPONENT 2: INTELLIGENCE — Scoring & Ranking Applicants
=========================================================
Takes applicants.json → outputs ranked_applicants.json + ranked_applicants.xlsx

Scoring model (total: 100 points):
  - Answer quality:       40 pts  (LLM-judged relevance + depth)
  - Technical skills:     20 pts  (keyword match + specificity)
  - GitHub quality:       20 pts  (real profile, activity, stars)
  - Anti-AI signals:      10 pts  (penalty-based, up to -10)
  - Completeness:         10 pts  (no blank fields)

Tiers:
  ≥80  → Fast-Track  (top 10%)
  60-79 → Standard
  40-59 → Review
  <40  → Reject
"""

import json
import re
import time
import logging
import argparse
import datetime
from pathlib import Path
from typing import Optional
import requests

log = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [C2-SCORE] %(levelname)s %(message)s"
)

INPUT_PATH  = Path("data/applicants.json")
OUTPUT_JSON = Path("data/ranked_applicants.json")
OUTPUT_XLSX = Path("data/ranked_applicants.xlsx")

# ── AI-GENERATION FINGERPRINTS ────────────────────────────────────────────────
# Phrases that appear disproportionately in LLM-generated text
AI_PHRASES = [
    "i'd be happy to",
    "i would be happy to",
    "here's a comprehensive",
    "comprehensive overview",
    "in today's rapidly evolving",
    "in the ever-evolving",
    "it's worth noting that",
    "it is worth noting",
    "at the end of the day",
    "in conclusion,",
    "to summarize,",
    "as an ai",
    "as a language model",
    "certainly! here",
    "absolutely! here",
    "great question!",
    "that's a great",
    "i hope this helps",
    "please let me know if",
    "feel free to reach out",
    "leverage",                 # overused in GPT outputs
    "utilize" ,                 # same
    "delve into",
    "dive deep into",
    "crucial role",
    "pivotal role",
    "it's important to note",
    "needless to say",
    "rest assured",
    "first and foremost",
    "last but not least",
    "furthermore,",
    "moreover,",
    "in addition to",
    "it's also worth",
    "when it comes to",
    "in terms of",
    "going forward",
    "moving forward",
]

# Technical skills we look for (domain: keywords)
TECHNICAL_SKILLS = {
    "python":       ["python", "django", "flask", "fastapi", "pandas", "numpy", "pytorch", "tensorflow"],
    "web":          ["javascript", "react", "vue", "angular", "node", "typescript", "html", "css", "nextjs"],
    "data":         ["sql", "postgresql", "mongodb", "redis", "elasticsearch", "kafka", "spark"],
    "devops":       ["docker", "kubernetes", "aws", "gcp", "azure", "ci/cd", "github actions", "terraform"],
    "scraping":     ["selenium", "playwright", "beautifulsoup", "scrapy", "requests", "puppeteer"],
    "ml":           ["machine learning", "deep learning", "nlp", "computer vision", "scikit-learn", "huggingface"],
    "system":       ["multithreading", "concurrency", "async", "microservices", "rest api", "graphql"],
}

# ── MAIN SCORER ───────────────────────────────────────────────────────────────

class ApplicantScorer:

    def __init__(self, use_llm: bool = True, anthropic_api_key: str = ""):
        self.use_llm = use_llm
        self.api_key = anthropic_api_key
        self._llm_cache = {}  # avoid re-scoring identical answers

    def score_all(self, applicants: list[dict]) -> list[dict]:
        """Score and rank all applicants. Returns sorted list with scores."""
        scored = []
        total = len(applicants)

        for i, app in enumerate(applicants):
            log.info(f"Scoring [{i+1}/{total}] {app.get('name', 'Unknown')}...")
            try:
                scored_app = self.score_one(app)
                scored.append(scored_app)
            except Exception as e:
                log.error(f"Failed to score {app.get('name')}: {e}")
                app["score"] = 0
                app["tier"] = "Reject"
                app["score_breakdown"] = {"error": str(e)}
                scored.append(app)

        # Sort by score descending
        scored.sort(key=lambda x: x.get("score", 0), reverse=True)

        # Add rank
        for rank, app in enumerate(scored, 1):
            app["rank"] = rank

        # Add percentile
        n = len(scored)
        for app in scored:
            app["percentile"] = round(100 * (1 - (app["rank"] - 1) / n), 1)

        log.info(f"\nScoring complete. Distribution:")
        tiers = {"Fast-Track": 0, "Standard": 0, "Review": 0, "Reject": 0}
        for app in scored:
            tiers[app.get("tier", "Reject")] += 1
        for tier, count in tiers.items():
            pct = 100 * count / n if n else 0
            log.info(f"  {tier:12s}: {count:4d} ({pct:.1f}%)")

        return scored

    def score_one(self, app: dict) -> dict:
        """Score a single applicant. Returns app dict with score fields added."""
        breakdown = {}

        # 1. Answer Quality (0-40)
        answer_score = self._score_answers(app, breakdown)

        # 2. Technical Skills (0-20)
        skills_score = self._score_skills(app, breakdown)

        # 3. GitHub Quality (0-20)
        github_score = self._score_github(app, breakdown)

        # 4. AI/Cheat Penalty (0 to -10)
        ai_penalty = self._compute_ai_penalty(app, breakdown)

        # 5. Completeness (0-10)
        completeness_score = self._score_completeness(app, breakdown)

        total = max(0, answer_score + skills_score + github_score + ai_penalty + completeness_score)
        total = min(100, total)

        # Tier assignment
        if total >= 80:
            tier = "Fast-Track"
        elif total >= 60:
            tier = "Standard"
        elif total >= 40:
            tier = "Review"
        else:
            tier = "Reject"

        app.update({
            "score": round(total, 1),
            "tier": tier,
            "score_breakdown": {
                "answer_quality":   round(answer_score, 1),
                "technical_skills": round(skills_score, 1),
                "github_quality":   round(github_score, 1),
                "ai_penalty":       round(ai_penalty, 1),
                "completeness":     round(completeness_score, 1),
                "total":            round(total, 1),
                "details":          breakdown
            }
        })
        return app

    # ── ANSWER QUALITY ─────────────────────────────────────────────────────────

    def _score_answers(self, app: dict, breakdown: dict) -> float:
        """
        Score answer quality. Uses LLM if available, falls back to heuristics.
        Max 40 points.
        """
        answers = app.get("answers", [])
        cover = app.get("cover_letter", "")
        all_text = cover + " " + " ".join(a.get("answer", "") for a in answers)
        all_text = all_text.strip()

        if not all_text:
            breakdown["answers"] = "no answers provided"
            return 0

        if self.use_llm and self.api_key:
            return self._score_answers_via_llm(app, breakdown)
        else:
            return self._score_answers_heuristic(all_text, answers, breakdown)

    def _score_answers_via_llm(self, app: dict, breakdown: dict) -> float:
        """Use Claude to score answer quality. Returns 0-40."""
        answers_text = "\n".join(
            f"Q: {a.get('question', 'General')}\nA: {a.get('answer', '')}"
            for a in app.get("answers", [])
        )
        if app.get("cover_letter"):
            answers_text = f"Cover letter: {app['cover_letter']}\n\n{answers_text}"

        cache_key = hash(answers_text[:500])
        if cache_key in self._llm_cache:
            return self._llm_cache[cache_key]

        prompt = f"""You are evaluating a job application for a software engineering internship.
Score this applicant's answers from 0 to 40 based on:
- Technical depth and accuracy (not just mentioning buzzwords)
- Specific examples and concrete details vs vague generalities
- Problem-solving approach shown
- Originality (penalize generic answers that could apply to any job)
- Language quality (penalize one-word answers, very short answers, or rambling)

APPLICANT ANSWERS:
{answers_text[:3000]}

Respond with ONLY a JSON object like:
{{"score": 28, "reasoning": "Shows solid Python knowledge with specific project..."}}

Do not explain. Just the JSON."""

        try:
            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": self.api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json"
                },
                json={
                    "model": "claude-sonnet-4-20250514",
                    "max_tokens": 200,
                    "messages": [{"role": "user", "content": prompt}]
                },
                timeout=20
            )
            data = resp.json()
            text = data["content"][0]["text"]
            result = json.loads(text)
            score = float(result.get("score", 0))
            score = max(0, min(40, score))
            breakdown["answers_llm"] = result.get("reasoning", "")
            self._llm_cache[cache_key] = score
            return score

        except Exception as e:
            log.warning(f"LLM scoring failed, falling back to heuristic: {e}")
            return self._score_answers_heuristic(
                app.get("cover_letter", ""), app.get("answers", []), breakdown
            )

    def _score_answers_heuristic(self, all_text: str, answers: list, breakdown: dict) -> float:
        """
        Heuristic answer scoring without LLM. Returns 0-40.
        """
        score = 0
        notes = []

        word_count = len(all_text.split())

        # Length signals engagement
        if word_count < 20:
            notes.append("very short answer (<20 words)")
            score += 3
        elif word_count < 50:
            notes.append("brief answer")
            score += 8
        elif word_count < 150:
            score += 18
        elif word_count < 400:
            score += 25
        else:
            score += 22  # very long can be padded — slight penalty

        # Specificity signals: numbers, proper nouns, code-like terms
        specificity_signals = re.findall(
            r'\b(?:\d+%|\d+x|\d+ months?|\d+ years?|[A-Z][a-z]+(?:JS|DB|SQL|API|ML))\b',
            all_text
        )
        score += min(8, len(specificity_signals) * 2)
        if specificity_signals:
            notes.append(f"specificity signals: {specificity_signals[:3]}")

        # Question relevance — check if they answered multiple questions
        if len(answers) > 1:
            answered = sum(1 for a in answers if len(a.get("answer", "").split()) > 5)
            if answered == len(answers):
                score += 5
                notes.append("answered all questions")
            elif answered < len(answers) * 0.5:
                score -= 5
                notes.append("skipped most questions")

        breakdown["answers_heuristic"] = "; ".join(notes)
        return min(40, max(0, score))

    # ── TECHNICAL SKILLS ───────────────────────────────────────────────────────

    def _score_skills(self, app: dict, breakdown: dict) -> float:
        """Score technical skill depth. Max 20 points."""
        all_text = (
            " ".join(app.get("skills", [])) + " " +
            app.get("cover_letter", "") + " " +
            " ".join(a.get("answer", "") for a in app.get("answers", []))
        ).lower()

        matched_domains = set()
        matched_skills = []

        for domain, keywords in TECHNICAL_SKILLS.items():
            for kw in keywords:
                if kw in all_text:
                    matched_domains.add(domain)
                    matched_skills.append(kw)

        # Breadth (different domains)
        breadth_score = min(10, len(matched_domains) * 2)

        # Depth (more keywords in same domain = knows it well)
        domain_counts = {}
        for domain, keywords in TECHNICAL_SKILLS.items():
            count = sum(1 for kw in keywords if kw in all_text)
            domain_counts[domain] = count
        max_domain_depth = max(domain_counts.values()) if domain_counts else 0
        depth_score = min(10, max_domain_depth * 2)

        total = breadth_score + depth_score
        breakdown["skills"] = f"domains: {matched_domains}, skills: {matched_skills[:5]}"
        return total

    # ── GITHUB QUALITY ─────────────────────────────────────────────────────────

    def _score_github(self, app: dict, breakdown: dict) -> float:
        """Score GitHub profile quality. Max 20 points."""
        github_info = app.get("github_info", {})

        if not github_info or not github_info.get("valid"):
            reason = github_info.get("reason", "no_url") if github_info else "no_url"
            breakdown["github"] = f"no valid profile ({reason})"
            return 0

        if github_info.get("is_empty"):
            breakdown["github"] = "empty profile (just created for the application)"
            return 2  # small credit for having a link, big penalty for empty

        score = 5  # base for having a real profile

        repos = github_info.get("repos", 0)
        stars = github_info.get("stars", 0)
        recent = github_info.get("recent_activity", False)
        non_forks = github_info.get("non_fork_count", 0)

        # Repos
        score += min(5, repos // 3)

        # Original work (not just forks)
        score += min(4, non_forks * 1)

        # Stars = others validated their work
        if stars >= 10:
            score += 4
        elif stars >= 3:
            score += 2
        elif stars >= 1:
            score += 1

        # Recent activity
        if recent:
            score += 2

        breakdown["github"] = (
            f"repos={repos}, stars={stars}, non-forks={non_forks}, "
            f"recent={recent}"
        )
        return min(20, score)

    # ── AI / CHEAT PENALTY ─────────────────────────────────────────────────────

    def _compute_ai_penalty(self, app: dict, breakdown: dict) -> float:
        """
        Returns a NEGATIVE number (penalty). Range: -10 to 0.
        High penalty = strong AI-generation signal.
        """
        all_text = (
            app.get("cover_letter", "") + " " +
            " ".join(a.get("answer", "") for a in app.get("answers", []))
        ).lower()

        if not all_text.strip():
            return 0

        matched_phrases = [p for p in AI_PHRASES if p in all_text]
        phrase_density = len(matched_phrases) / max(1, len(all_text.split()) / 100)

        # Structural AI signals
        structural_signals = []

        # Perfect paragraph structure (AI loves 3-paragraph answers)
        paragraphs = [p.strip() for p in all_text.split("\n\n") if p.strip()]
        if len(paragraphs) == 3 and all(len(p.split()) > 30 for p in paragraphs):
            structural_signals.append("3-paragraph structure")

        # Bullet lists with uniform length (AI trait)
        bullet_items = re.findall(r'[-•]\s+(.+)', all_text)
        if len(bullet_items) >= 3:
            lengths = [len(item.split()) for item in bullet_items]
            avg = sum(lengths) / len(lengths)
            variance = sum((l - avg) ** 2 for l in lengths) / len(lengths)
            if variance < 10:  # very uniform bullet points
                structural_signals.append("uniform bullet length")

        # Count total signals
        total_signals = len(matched_phrases) + len(structural_signals)

        if total_signals == 0:
            penalty = 0
            note = "no AI signals detected"
        elif total_signals <= 2:
            penalty = -2
            note = f"mild AI signals: {matched_phrases[:2]}"
        elif total_signals <= 5:
            penalty = -5
            note = f"moderate AI signals: {matched_phrases[:3]} + {structural_signals}"
        else:
            penalty = -10
            note = f"strong AI signals ({total_signals} total): {matched_phrases[:5]}"

        breakdown["ai_penalty"] = note
        app["ai_flags"] = matched_phrases + structural_signals  # for anti-cheat component
        return penalty

    # ── COMPLETENESS ───────────────────────────────────────────────────────────

    def _score_completeness(self, app: dict, breakdown: dict) -> float:
        """Score how complete the profile is. Max 10 points."""
        score = 0
        missing = []

        if app.get("email"):
            score += 2
        else:
            missing.append("email")

        if app.get("cover_letter", "").strip():
            score += 2
        else:
            missing.append("cover letter")

        if app.get("answers"):
            score += 2
        else:
            missing.append("screening answers")

        if app.get("github_url"):
            score += 2
        else:
            missing.append("GitHub URL")

        if app.get("skills"):
            score += 2
        else:
            missing.append("skills")

        if missing:
            breakdown["completeness"] = f"missing: {', '.join(missing)}"
        else:
            breakdown["completeness"] = "all fields present"

        return score


# ── XLSX EXPORT ───────────────────────────────────────────────────────────────

def export_to_xlsx(ranked: list[dict], output_path: Path):
    """Export ranked applicants to a formatted Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("Install openpyxl: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Applicants"

    # Colors for tiers
    tier_fills = {
        "Fast-Track": PatternFill("solid", fgColor="C6EFCE"),   # green
        "Standard":   PatternFill("solid", fgColor="FFEB9C"),   # yellow
        "Review":     PatternFill("solid", fgColor="FFCC99"),   # orange
        "Reject":     PatternFill("solid", fgColor="FFC7CE"),   # red
    }

    headers = [
        "Rank", "Name", "Email", "Score", "Tier", "Percentile",
        "Answer Quality", "Tech Skills", "GitHub", "AI Penalty", "Completeness",
        "GitHub URL", "Skills", "AI Flags", "Cover Letter (preview)"
    ]

    # Header row
    header_fill = PatternFill("solid", fgColor="2E75B6")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, app in enumerate(ranked, 2):
        breakdown = app.get("score_breakdown", {})
        tier = app.get("tier", "Reject")
        row_fill = tier_fills.get(tier, PatternFill("solid", fgColor="FFFFFF"))

        row_data = [
            app.get("rank"),
            app.get("name"),
            app.get("email"),
            app.get("score"),
            tier,
            app.get("percentile"),
            breakdown.get("answer_quality"),
            breakdown.get("technical_skills"),
            breakdown.get("github_quality"),
            breakdown.get("ai_penalty"),
            breakdown.get("completeness"),
            app.get("github_url", ""),
            ", ".join(app.get("skills", [])[:5]),
            ", ".join(app.get("ai_flags", [])[:3]),
            (app.get("cover_letter", "") or "")[:200],
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill

    # Column widths
    col_widths = [6, 20, 25, 8, 12, 10, 14, 12, 10, 10, 12, 35, 30, 30, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    tiers = {"Fast-Track": 0, "Standard": 0, "Review": 0, "Reject": 0}
    for app in ranked:
        tiers[app.get("tier", "Reject")] += 1
    n = len(ranked)

    ws2["A1"] = "Tier"
    ws2["B1"] = "Count"
    ws2["C1"] = "Percentage"
    for i, (tier, count) in enumerate(tiers.items(), 2):
        ws2.cell(row=i, column=1, value=tier)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=f"{100*count/n:.1f}%")

    wb.save(output_path)
    log.info(f"Excel file saved: {output_path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Score and rank applicants")
    parser.add_argument("--input", default=str(INPUT_PATH), help="Input JSON file")
    parser.add_argument("--no-llm", action="store_true", help="Use heuristics only (no API call)")
    parser.add_argument("--api-key", default="", help="Anthropic API key for LLM scoring")
    parser.add_argument("--top-n", type=int, help="Print top N candidates")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        log.error(f"Input file not found: {input_path}")
        return

    with open(input_path) as f:
        applicants = json.load(f)

    log.info(f"Loaded {len(applicants)} applicants from {input_path}")

    use_llm = not args.no_llm and bool(args.api_key)
    if use_llm:
        log.info("LLM scoring enabled (Claude API)")
    else:
        log.info("Using heuristic scoring (fast, no API)")

    scorer = ApplicantScorer(use_llm=use_llm, anthropic_api_key=args.api_key)
    ranked = scorer.score_all(applicants)

    # Save JSON
    with open(OUTPUT_JSON, "w") as f:
        json.dump(ranked, f, indent=2, ensure_ascii=False)
    log.info(f"Ranked JSON saved: {OUTPUT_JSON}")

    # Save Excel
    export_to_xlsx(ranked, OUTPUT_XLSX)

    # Print top N
    if args.top_n:
        print(f"\n{'='*60}")
        print(f"TOP {args.top_n} CANDIDATES")
        print(f"{'='*60}")
        for app in ranked[:args.top_n]:
            print(
                f"#{app['rank']:3d} | {app['score']:5.1f} | {app['tier']:12s} | "
                f"{app.get('name', 'N/A'):25s} | {app.get('email', '')}"
            )


if __name__ == "__main__":
    main()
